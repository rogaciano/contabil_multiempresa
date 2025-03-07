from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.db.models import Q
from django.contrib.auth.models import User
from django.utils.translation import gettext as _
import csv
import json

from .models import Account, AccountType, UserProfile, Company
from transactions.models import Transaction
from .forms import AccountForm, UserProfileForm, UserForm

class AccountListView(LoginRequiredMixin, ListView):
    model = Account
    template_name = 'accounts/account_list.html'
    context_object_name = 'accounts'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Obter a empresa atual do usuário
        current_company_id = self.request.session.get('current_company_id')
        if not current_company_id:
            # Se não houver empresa selecionada, retornar queryset vazio
            messages.warning(self.request, _('Selecione uma empresa para visualizar as contas.'))
            return Account.objects.none()
            
        # Filtrar contas pela empresa atual
        queryset = queryset.filter(company_id=current_company_id)
        
        # Filtro por tipo de conta
        account_type = self.request.GET.get('type')
        if account_type:
            queryset = queryset.filter(type=account_type)
        
        # Pesquisa por código ou nome
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(code__icontains=search_query) | 
                Q(name__icontains=search_query)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['account_types'] = AccountType.choices
        context['search_query'] = self.request.GET.get('search', '')
        return context

class AccountDetailView(LoginRequiredMixin, DetailView):
    model = Account
    template_name = 'accounts/account_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account = self.get_object()
        
        # Obter transações da conta
        context['transactions'] = account.debit_transactions.all() | account.credit_transactions.all()
        context['balance'] = account.get_balance()
        
        return context

class AccountCreateView(LoginRequiredMixin, CreateView):
    model = Account
    template_name = 'accounts/account_form.html'
    form_class = AccountForm
    success_url = reverse_lazy('account_list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Passar o ID da empresa atual para o formulário
        current_company_id = self.request.session.get('current_company_id')
        if current_company_id:
            kwargs['company_id'] = current_company_id
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Adicionar a empresa atual ao contexto
        current_company_id = self.request.session.get('current_company_id')
        context['current_company_id'] = current_company_id
        
        # Verificar se o usuário selecionou uma empresa
        if not current_company_id:
            messages.warning(self.request, _('Selecione uma empresa antes de criar uma conta.'))
        
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        
        # Verificar se o usuário selecionou uma empresa
        current_company_id = self.request.session.get('current_company_id')
        if not current_company_id:
            messages.error(self.request, _('Selecione uma empresa para criar uma conta.'))
            return self.form_invalid(form)
        
        # Definir a empresa da conta antes da validação
        from .models import Company
        try:
            company = Company.objects.get(id=current_company_id)
            form.instance.company = company
            
            if form.is_valid():
                try:
                    return self.form_valid(form)
                except Exception as e:
                    # Capturar erro de unicidade (código de conta duplicado)
                    from django.db import IntegrityError
                    if isinstance(e, IntegrityError) and 'UNIQUE constraint failed: accounts_account.company_id, accounts_account.code' in str(e):
                        form.add_error('code', _('Já existe uma conta com este código nesta empresa. Por favor, escolha um código diferente.'))
                        messages.error(self.request, _('Erro ao criar conta: código já existe nesta empresa.'))
                        return self.form_invalid(form)
                    else:
                        # Repassar outros erros
                        raise
            else:
                return self.form_invalid(form)
        except Company.DoesNotExist:
            messages.error(self.request, _('Empresa não encontrada.'))
            return self.form_invalid(form)
    
    def form_valid(self, form):
        messages.success(self.request, _('Conta criada com sucesso!'))
        return super().form_valid(form)

class AccountUpdateView(LoginRequiredMixin, UpdateView):
    model = Account
    template_name = 'accounts/account_form.html'
    form_class = AccountForm
    success_url = reverse_lazy('account_list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Passar o ID da empresa atual para o formulário
        current_company_id = self.request.session.get('current_company_id')
        if current_company_id:
            kwargs['company_id'] = current_company_id
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Adicionar a empresa atual ao contexto
        current_company_id = self.request.session.get('current_company_id')
        context['current_company_id'] = current_company_id
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        
        # Verificar se a empresa da conta sendo editada corresponde à empresa atual
        account = self.object
        current_company_id = self.request.session.get('current_company_id')
        
        if current_company_id and str(account.company_id) != str(current_company_id):
            messages.error(self.request, _('Você só pode editar contas da empresa atual.'))
            return self.form_invalid(form)
        
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)
    
    def form_valid(self, form):
        try:
            messages.success(self.request, _('Conta atualizada com sucesso!'))
            return super().form_valid(form)
        except Exception as e:
            # Capturar erro de unicidade (código de conta duplicado)
            from django.db import IntegrityError
            if isinstance(e, IntegrityError) and 'UNIQUE constraint failed: accounts_account.company_id, accounts_account.code' in str(e):
                form.add_error('code', _('Já existe uma conta com este código nesta empresa. Por favor, escolha um código diferente.'))
                messages.error(self.request, _('Erro ao atualizar conta: código já existe nesta empresa.'))
                return self.form_invalid(form)
            else:
                # Repassar outros erros
                raise

class AccountDeleteView(LoginRequiredMixin, DeleteView):
    model = Account
    template_name = 'accounts/account_confirm_delete.html'
    success_url = reverse_lazy('account_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        account = self.get_object()
        
        # Verificar se existem transações associadas à conta
        transactions_count = Transaction.objects.filter(
            Q(debit_account=account) | Q(credit_account=account)
        ).count()
        
        context['transactions_count'] = transactions_count
        return context
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Conta excluída com sucesso!'))
        return super().delete(request, *args, **kwargs)

class AccountChartView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/account_chart.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obter a empresa atual do usuário
        current_company_id = self.request.session.get('current_company_id')
        if not current_company_id:
            messages.warning(self.request, _('Selecione uma empresa para visualizar o plano de contas.'))
            context['account_tree'] = {'A': [], 'L': [], 'E': [], 'R': [], 'X': []}
            return context
        
        # Criar estrutura hierárquica das contas da empresa atual
        accounts = Account.objects.filter(company_id=current_company_id)
        
        # Organizar contas por tipo
        account_tree = {
            'A': [],  # Ativos
            'L': [],  # Passivos
            'E': [],  # Patrimônio Líquido
            'R': [],  # Receitas
            'X': []   # Despesas
        }
        
        # Mapear todas as contas por ID para facilitar a construção da árvore
        accounts_by_id = {account.id: {
            'id': account.id,
            'code': account.code,
            'name': account.name,
            'type': account.type,
            'type_display': account.get_type_display(),
            'children': []
        } for account in accounts}
        
        # Construir a estrutura hierárquica
        root_accounts = []
        for account in accounts:
            account_data = accounts_by_id[account.id]
            
            if account.parent:
                # Adicionar como filho da conta pai
                parent_data = accounts_by_id.get(account.parent.id)
                if parent_data:
                    parent_data['children'].append(account_data)
            else:
                # Conta de nível raiz
                root_accounts.append(account_data)
                account_tree[account.type].append(account_data)
        
        context['account_tree'] = account_tree
        return context

class AccountImportView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/account_import.html'
    
    def post(self, request, *args, **kwargs):
        csv_file = request.FILES.get('csv_file')
        update_existing = request.POST.get('update_existing') == 'on'
        
        if not csv_file or not csv_file.name.endswith('.csv'):
            messages.error(request, _('Por favor, selecione um arquivo CSV válido.'))
            return self.get(request, *args, **kwargs)
        
        # Obter a empresa atual do usuário
        current_company_id = request.session.get('current_company_id')
        if not current_company_id:
            messages.error(request, _('Selecione uma empresa antes de importar contas.'))
            return self.get(request, *args, **kwargs)
        
        results = {
            'total_rows': 0,
            'created': 0,
            'updated': 0,
            'errors': 0,
            'error_details': []
        }
        
        try:
            # Decodificar o arquivo CSV
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            
            # Mapear códigos de contas para IDs para facilitar a associação de contas pai
            # Filtrar apenas contas da empresa atual
            code_to_id = {account.code: account.id for account in Account.objects.filter(company_id=current_company_id)}
            
            for row in reader:
                results['total_rows'] += 1
                
                try:
                    code = row.get('code', '').strip()
                    name = row.get('name', '').strip()
                    type_code = row.get('type', '').strip().upper()
                    parent_code = row.get('parent_code', '').strip()
                    description = row.get('description', '').strip()
                    is_active = row.get('is_active', 'true').lower() in ('true', 'yes', '1')
                    
                    # Validar campos obrigatórios
                    if not code or not name or not type_code:
                        raise ValueError(_('Código, nome e tipo são campos obrigatórios.'))
                    
                    # Validar tipo de conta
                    if type_code not in [choice[0] for choice in AccountType.choices]:
                        raise ValueError(_('Tipo de conta inválido: ') + type_code)
                    
                    # Verificar se a conta já existe na empresa atual
                    account = Account.objects.filter(code=code, company_id=current_company_id).first()
                    
                    if account and update_existing:
                        # Atualizar conta existente
                        account.name = name
                        account.type = type_code
                        account.description = description
                        account.is_active = is_active
                        
                        # Atualizar conta pai se fornecida
                        if parent_code:
                            parent_id = code_to_id.get(parent_code)
                            if parent_id:
                                account.parent_id = parent_id
                            else:
                                raise ValueError(_('Conta pai não encontrada: ') + parent_code)
                        else:
                            account.parent = None
                        
                        account.save()
                        results['updated'] += 1
                        
                    elif not account:
                        # Criar nova conta
                        parent_id = None
                        if parent_code:
                            parent_id = code_to_id.get(parent_code)
                            if not parent_id:
                                raise ValueError(_('Conta pai não encontrada: ') + parent_code)
                        
                        account = Account.objects.create(
                            code=code,
                            name=name,
                            type=type_code,
                            parent_id=parent_id,
                            description=description,
                            is_active=is_active,
                            company_id=current_company_id  # Associar à empresa atual
                        )
                        
                        # Atualizar o mapeamento de códigos para IDs
                        code_to_id[code] = account.id
                        results['created'] += 1
                    
                    else:
                        # Conta existe mas não deve ser atualizada
                        results['error_details'].append(_('Linha ') + str(results["total_rows"]) + _(': Conta ') + code + _(' já existe e a opção de atualização não está ativada.'))
                        results['errors'] += 1
                
                except Exception as e:
                    results['error_details'].append(_('Linha ') + str(results["total_rows"]) + _(': ') + str(e))
                    results['errors'] += 1
            
            if results['created'] > 0 or results['updated'] > 0:
                messages.success(
                    request, 
                    _('Importação concluída: ') + str(results["created"]) + _(' contas criadas, ') + str(results["updated"]) + _(' contas atualizadas.')
                )
            
            if results['errors'] > 0:
                messages.warning(
                    request, 
                    _('Importação concluída com ') + str(results["errors"]) + _(' erros. Verifique os detalhes abaixo.')
                )
        
        except Exception as e:
            messages.error(request, _('Erro ao processar o arquivo: ') + str(e))
            results['errors'] += 1
            results['error_details'].append(str(e))
        
        context = self.get_context_data(**kwargs)
        context['results'] = results
        return self.render_to_response(context)

class AccountsByTypeView(LoginRequiredMixin, View):
    def get(self, request, type_code, *args, **kwargs):
        if type_code not in [choice[0] for choice in AccountType.choices]:
            return JsonResponse({'error': _('Tipo de conta inválido')}, status=400)
        
        # Obter a empresa atual do usuário
        current_company_id = request.session.get('current_company_id')
        if not current_company_id:
            return JsonResponse({'error': _('Nenhuma empresa selecionada')}, status=400)
        
        accounts = Account.objects.filter(
            type=type_code,
            company_id=current_company_id
        ).values('id', 'code', 'name')
        
        return JsonResponse(list(accounts), safe=False)

class AccountExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        account_type = request.GET.get('type')
        active_only = request.GET.get('active_only') == '1'
        include_balance = request.GET.get('include_balance') == '1'
        export_format = request.GET.get('format', 'csv')
        
        # Obter a empresa atual do usuário
        current_company_id = request.session.get('current_company_id')
        if not current_company_id:
            messages.error(request, _('Selecione uma empresa antes de exportar contas.'))
            return redirect('account_list')
        
        # Filtrar contas pela empresa atual
        queryset = Account.objects.filter(company_id=current_company_id)
        
        if account_type:
            queryset = queryset.filter(type=account_type)
        
        if active_only:
            queryset = queryset.filter(is_active=True)
        
        # Ordenar por código
        queryset = queryset.order_by('code')
        
        if export_format == 'excel':
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font, Alignment, PatternFill
                
                # Criar um novo workbook e selecionar a planilha ativa
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = _("Plano de Contas")
                
                # Definir cabeçalhos
                headers = [_('Código'), _('Nome'), _('Tipo'), _('Conta Pai'), _('Descrição'), _('Ativo')]
                if include_balance:
                    headers.append(_('Saldo'))
                
                # Estilo para cabeçalhos
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                
                # Adicionar cabeçalhos
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Adicionar dados
                for row_num, account in enumerate(queryset, 2):
                    ws.cell(row=row_num, column=1, value=account.code)
                    ws.cell(row=row_num, column=2, value=account.name)
                    ws.cell(row=row_num, column=3, value=account.get_type_display())
                    ws.cell(row=row_num, column=4, value=account.parent.code if account.parent else '')
                    ws.cell(row=row_num, column=5, value=account.description)
                    ws.cell(row=row_num, column=6, value=_('Sim') if account.is_active else _('Não'))
                    
                    if include_balance:
                        ws.cell(row=row_num, column=7, value=float(account.get_balance()))
                
                # Ajustar largura das colunas
                for col_num in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_num)
                    ws.column_dimensions[column_letter].width = 15
                
                # Configurar a resposta HTTP
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="plano_de_contas.xlsx"'
                
                # Salvar o workbook para a resposta
                wb.save(response)
                return response
                
            except ImportError:
                messages.warning(request, _('Pacote openpyxl não está instalado. Exportando como CSV.'))
                export_format = 'csv'
        
        # Exportar como CSV (padrão)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="plano_de_contas.csv"'
        
        writer = csv.writer(response)
        
        # Escrever cabeçalhos
        headers = ['code', 'name', 'type', 'parent_code', 'description', 'is_active']
        if include_balance:
            headers.append('balance')
        
        writer.writerow(headers)
        
        # Escrever dados
        for account in queryset:
            row = [
                account.code,
                account.name,
                account.type,
                account.parent.code if account.parent else '',
                account.description,
                'true' if account.is_active else 'false'
            ]
            
            if include_balance:
                row.append(float(account.get_balance()))
            
            writer.writerow(row)
        
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['account_types'] = AccountType.choices
        return context

class AccountUpdateFormView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/account_form_update.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['account_types'] = AccountType.choices
        return context

class UserProfileView(LoginRequiredMixin, DetailView):
    model = UserProfile
    template_name = 'accounts/user_profile.html'
    context_object_name = 'profile'
    
    def get_object(self, queryset=None):
        # Retorna o perfil do usuário atual
        return self.request.user.profile

class UserProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = UserProfile
    form_class = UserProfileForm
    template_name = 'accounts/user_profile_form.html'
    success_url = reverse_lazy('user_profile')
    
    def get_object(self, queryset=None):
        # Retorna o perfil do usuário atual
        return self.request.user.profile
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['user_form'] = UserForm(self.request.POST, instance=self.request.user)
        else:
            context['user_form'] = UserForm(instance=self.request.user)
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        user_form = context['user_form']
        if user_form.is_valid():
            user_form.save()
            form.save()
            messages.success(self.request, _('Perfil e usuário atualizados com sucesso!'))
            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

class UserSettingsView(LoginRequiredMixin, TemplateView):
    template_name = 'accounts/user_settings.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = self.request.user
        return context

class AccountTemplateDownloadView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="account_import_template.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['code', 'name', 'type', 'parent_code', 'description', 'is_active'])
        writer.writerow(['1', 'Ativo', 'A', '', 'Contas de Ativo', 'true'])
        writer.writerow(['1.1', 'Ativo Circulante', 'A', '1', 'Contas de Ativo Circulante', 'true'])
        writer.writerow(['1.1.1', 'Caixa', 'A', '1.1', 'Dinheiro em espécie', 'true'])
        writer.writerow(['2', 'Passivo', 'L', '', 'Contas de Passivo', 'true'])
        writer.writerow(['2.1', 'Passivo Circulante', 'L', '2', 'Contas de Passivo Circulante', 'true'])
        writer.writerow(['3', 'Patrimônio Líquido', 'E', '', 'Contas de Patrimônio Líquido', 'true'])
        writer.writerow(['4', 'Receitas', 'R', '', 'Contas de Receita', 'true'])
        writer.writerow(['5', 'Despesas', 'X', '', 'Contas de Despesa', 'true'])
        
        return response
